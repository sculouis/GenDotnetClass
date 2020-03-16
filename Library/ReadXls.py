from openpyxl import load_workbook
from collections import namedtuple

class ReadXls:
    def __init__(self,config):
        #data_only=True means formula is calculate result
        self.filename = config.get('filename')
        self.wb = load_workbook(filename = self.filename,data_only=True) 

    def fieldTitle(self,sheetName,rowPoisition):
        """get field Name"""
        ws = self.wb[sheetName]
        rowObj = []
        for row in ws.iter_rows(min_row=rowPoisition, max_col=ws.max_column, max_row=rowPoisition): 
            rowObj = ([row[i].value for i in range(ws.max_column) if row[i].value != None])
        return rowObj

    def getSheetData(self,sheetName,startRow = 3,endcol = 0):
        """get data by specified sheetname"""
        ws = self.wb[sheetName]
        rows = []
        for row in ws.iter_rows(min_row=startRow, max_col=endcol, max_row=ws.max_row): 
            if row[0].value == None:
                print('this row is None')
            else:
                rows.append([row[i] for i in range(endcol)])
        return rows
        # for row in rows: 
        #     for cell in row:
        #         print(cell.value)

    def getSheetNames(self,NotIn = ['Index'] ):
        """get all sheet Names"""
        return [sheetName for sheetName in self.wb.sheetnames if sheetName not in NotIn]

    def GetRows(self,sheetName,startRow = 4):
        titles = self.fieldTitle(sheetName,3)   #取得欄位名稱    
        colLen = len(titles)
        rows = self.getSheetData(sheetName,startRow,colLen)
        mapRows = []
        for row in rows: 
            Fields=namedtuple("Fields",titles)
            myList = [field.value for field in row]
            fields = Fields(*myList)
            mapRows.append(fields)
        return mapRows